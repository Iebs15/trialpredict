import numpy as np
from openpyxl import load_workbook
from typing import List, Tuple, Dict, Any
import matplotlib.pyplot as plt
from matplotlib.patches import Patch

# ----------------------------------------------------------
# Utility: Convert openpyxl Color or ARGB/RGB string → 6-digit RGB hex
# ----------------------------------------------------------
def rgb_to_hex(rgb) -> str:
    """
    Convert an openpyxl Color object or ARGB/RGB hex string into a 6-digit RGB hex string.
    Examples:
      - "FF00FF00" → "00FF00"
      - "FFFFFF00" → "FFFF00"
      - "C6EFCE"   → "C6EFCE"
    """
    if rgb is None:
        return None

    # If we got a raw string like "FF00FF00" or "C6EFCE"
    if isinstance(rgb, str):
        s = rgb.upper()
        if len(s) == 8:  # ARGB
            return s[-6:]
        if len(s) == 6:  # already RGB
            return s
        return None

    # Otherwise, assume it's a Color object with an .rgb attribute
    try:
        raw = rgb.rgb  # e.g. "FF00FF00"
    except AttributeError:
        return None

    if raw is None:
        return None
    raw = raw.upper()
    if len(raw) == 8:  # ARGB
        return raw[-6:]
    if len(raw) == 6:  # RGB
        return raw
    return None

# ----------------------------------------------------------
# Utility: Classify cell.fill color → "inhibitor"/"promoter"/"unknown"/None
# ----------------------------------------------------------
def classify_fill(cell) -> str:
    """
    Inspect cell.fill.fgColor.rgb (or indexed color). Return:
      - "inhibitor" if hex == "C6EFCE" (green).
      - "promoter"  if hex == "FFEB9C" (yellow).
      - "unknown"   if the cell.value is numeric but hex is neither.
      - None otherwise (text, blank, non-numeric).
    """
    fill = cell.fill
    hexcode = None
    if fill is not None:
        fg = fill.fgColor
        hexcode = rgb_to_hex(fg)

    if hexcode == "C6EFCE":
        return "inhibitor"
    if hexcode == "FFEB9C":
        return "promoter"

    val = cell.value
    if isinstance(val, (int, float)):
        return "unknown"

    return None

# ----------------------------------------------------------
# Build weighted association from Excel
# ----------------------------------------------------------
def build_weighted_association(wb, symptom: str) -> Dict[str, Dict[str, Dict[str, Any]]]:
    """
    Reads the Excel file and, for each biomarker–disease pair (where `unique_symptom` == symptom),
    collects numeric scores separated into three categories: inhibitor, promoter, unknown.
    Then computes average & percentage breakdowns. Returns a nested dict:
      { biomarker: { disease: { "avg_inhibitor": ..., "percent_inhibitor": ..., 
                               "avg_promoter":  ..., "percent_promoter":  ..., 
                               "avg_unknown":   ..., "percent_unknown":   ..., 
                               "total_avg":      ... } } }
    Prunes out any (biomarker, disease) if all four stats are None.
    """
    # wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active

    # 1) Find column indexes for "unique_symptom" and "MappedDisease"
    header = [cell.value for cell in sheet[1]]
    sym_idx = None
    disease_idx = None
    for i, h in enumerate(header):
        if isinstance(h, str):
            low = h.strip().lower()
            if low == "unique_symptom":
                sym_idx = i
            if low == "mappeddisease":
                disease_idx = i

    if sym_idx is None:
        raise ValueError("Could not find 'unique_symptom' in header.")
    if disease_idx is None:
        raise ValueError("Could not find 'MappedDisease' in header.")

    # 2) All columns to the right of unique_symptom are biomarker names
    biomarker_names = header[sym_idx + 1:]
    if not biomarker_names:
        raise ValueError("No biomarker columns found to the right of 'unique_symptom'.")

    # Prepare raw-score buckets
    inh_dict = {}
    prm_dict = {}
    unk_dict = {}

    matched_count = 0
    # 3) Iterate data rows (2..max_row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        sym_cell = row[sym_idx]
        if sym_cell.value is None:
            continue
        if str(sym_cell.value).strip().lower() != symptom.lower():
            continue

        matched_count += 1
        disease_cell = row[disease_idx]
        if disease_cell.value is None:
            continue
        disease_name = str(disease_cell.value).strip()

        for offset, biomarker in enumerate(biomarker_names, start=sym_idx + 1):
            cell = row[offset]
            val = cell.value
            if not isinstance(val, (int, float)):
                continue
            category = classify_fill(cell)
            key = (biomarker, disease_name)
            if category == "inhibitor":
                inh_dict.setdefault(key, []).append(float(val))
            elif category == "promoter":
                prm_dict.setdefault(key, []).append(float(val))
            elif category == "unknown":
                unk_dict.setdefault(key, []).append(float(val))

    # 4) Build nested result, prune all-None entries
    result = {}
    for biomarker in biomarker_names:
        kept = {}
        # gather diseases for which this biomarker has any scores
        disease_set = set(
            d for (b, d) in list(inh_dict.keys()) + list(prm_dict.keys()) + list(unk_dict.keys())
            if b == biomarker
        )
        for disease in disease_set:
            key = (biomarker, disease)
            inh_list = inh_dict.get(key, [])
            prm_list = prm_dict.get(key, [])
            unk_list = unk_dict.get(key, [])

            avg_inh = float(np.mean(inh_list)) if inh_list else None
            avg_prm = float(np.mean(prm_list)) if prm_list else None
            avg_unk = float(np.mean(unk_list)) if unk_list else None

            total_scores = inh_list + prm_list + unk_list
            total_avg = float(np.mean(total_scores)) if total_scores else None

            # prune if all four are None
            if not any(v is not None for v in (avg_inh, avg_prm, avg_unk, total_avg)):
                continue

            # compute percentages (treat None as 0)
            part_inh = avg_inh if avg_inh is not None else 0.0
            part_prm = avg_prm if avg_prm is not None else 0.0
            part_unk = avg_unk if avg_unk is not None else 0.0
            sum_parts = part_inh + part_prm + part_unk
            if sum_parts > 0:
                pct_inh = round((part_inh / sum_parts) * 100, 1)
                pct_prm = round((part_prm / sum_parts) * 100, 1)
                pct_unk = round((part_unk / sum_parts) * 100, 1)
            else:
                pct_inh = pct_prm = pct_unk = 0.0

            kept[disease] = {
                "avg_inhibitor":     avg_inh,
                "percent_inhibitor": pct_inh,
                "avg_promoter":      avg_prm,
                "percent_promoter":  pct_prm,
                "avg_unknown":       avg_unk,
                "percent_unknown":   pct_unk,
                "total_avg":         total_avg,
            }

        if kept:
            result[biomarker] = kept

    print(f"Found {matched_count} row(s) matching symptom '{symptom}'.")
    return result

# ----------------------------------------------------------
# Utility: Collect raw averages by color
# ----------------------------------------------------------
def collect_biomarker_scores_by_color(wb, symptom: str) -> Tuple[List[str], Dict[str, Dict[str, float]], Dict[str, Dict[str, float]], Dict[str, Dict[str, float]]]:
    """
    Read Excel, find rows where `unique_symptom`==symptom, then for each biomarker column:
    bucket numeric values into inhibitor/promoter/unknown by color fill. Return:
      - biomarker_names: list[str]
      - inhibitor_avg:    dict[str → (disease → avg_score)]
      - promoter_avg:     dict[str → (disease → avg_score)]
      - unknown_avg:      dict[str → (disease → avg_score)]
    """
    # wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active

    # 1) Header row (row 1)
    header_raw = [cell.value for cell in sheet[1]]
    sym_col_idx = None
    disease_col_idx = None
    for i, h in enumerate(header_raw):
        if isinstance(h, str):
            if h.strip().lower() == "unique_symptom":
                sym_col_idx = i
            if h.strip().lower() == "mappeddisease":
                disease_col_idx = i

    if sym_col_idx is None:
        raise ValueError("Could not find column 'unique_symptom' in header.")
    if disease_col_idx is None:
        raise ValueError("Could not find column 'MappedDisease' in header.")

    # 2) Biomarker columns to the right of unique_symptom
    biomarker_names = header_raw[sym_col_idx + 1:]
    if not biomarker_names:
        raise ValueError("No columns found to the right of 'unique_symptom'—nothing to process.")

    # 3) Prepare raw lists: biomarker → disease → [scores]
    inh_dict = {b: {} for b in biomarker_names}
    prm_dict = {b: {} for b in biomarker_names}
    unk_dict = {b: {} for b in biomarker_names}

    matched_rows = 0
    # 4) Iterate data rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        sym_cell = row[sym_col_idx]
        if sym_cell.value is None:
            continue
        if str(sym_cell.value).strip().lower() != symptom.lower():
            continue

        matched_rows += 1
        disease_name = str(row[disease_col_idx].value).strip() if row[disease_col_idx].value else "UnknownDisease"

        for offset, bio in enumerate(biomarker_names, start=sym_col_idx + 1):
            cell = row[offset]
            val = cell.value
            if not isinstance(val, (int, float)):
                continue
            category = classify_fill(cell)
            if category == "inhibitor":
                inh_dict[bio].setdefault(disease_name, []).append(float(val))
            elif category == "promoter":
                prm_dict[bio].setdefault(disease_name, []).append(float(val))
            elif category == "unknown":
                unk_dict[bio].setdefault(disease_name, []).append(float(val))

    print(f"Found {matched_rows} row(s) matching symptom '{symptom}'.")

    # 5) Convert raw lists → average per disease
    inhibitor_avg = {}
    promoter_avg  = {}
    unknown_avg   = {}

    for bio in biomarker_names:
        inh_avg = {d: float(np.mean(scores)) for d, scores in inh_dict[bio].items() if scores}
        prm_avg = {d: float(np.mean(scores)) for d, scores in prm_dict[bio].items() if scores}
        unk_avg = {d: float(np.mean(scores)) for d, scores in unk_dict[bio].items() if scores}

        if inh_avg or prm_avg or unk_avg:
            inhibitor_avg[bio] = inh_avg
            promoter_avg[bio]  = prm_avg
            unknown_avg[bio]   = unk_avg

    return biomarker_names, inhibitor_avg, promoter_avg, unknown_avg

# ----------------------------------------------------------
# Plotting helper
# ----------------------------------------------------------
def plot_stacked_bars_per_disease(biomarker_names: List[str], category_avg: Dict[str, Dict[str, float]], symptom: str, category_name: str, colormap_name: str = "tab20") -> None:
    """
    Draw a stacked-bar chart: each bar = one biomarker, each slice = one disease’s average score.
    """
    # 1) Filter out biomarkers with no data
    filtered_bio = [b for b in biomarker_names if b in category_avg and category_avg[b]]
    if not filtered_bio:
        print(f"No {category_name.lower()} data found for '{symptom}'. Skipping plot.")
        return

    # 2) Gather all diseases across filtered biomarkers
    all_diseases = sorted({d for b in filtered_bio for d in category_avg[b].keys()})
    num_diseases = len(all_diseases)

    # 3) Build color map
    cmap = plt.get_cmap(colormap_name)
    disease_to_color = {d: cmap(i % cmap.N) for i, d in enumerate(all_diseases)}

    # 4) X positions
    N = len(filtered_bio)
    x = np.arange(N)
    width = 0.8

    fig, ax = plt.subplots(figsize=(14, 6))

    # 5) For each biomarker, stack disease slices (sorted ascending by score)
    for i, bio in enumerate(filtered_bio):
        bottom = 0.0
        disease_scores = list(category_avg[bio].items())
        disease_scores.sort(key=lambda it: it[1])  # ascending by avg_score

        for disease, avg_score in disease_scores:
            if avg_score <= 0:
                continue
            color = disease_to_color[disease]

            ax.bar(
                x[i],
                avg_score,
                bottom=bottom,
                color=color,
                width=width,
                edgecolor="white"
            )
            # Label the slice
            mid_height = bottom + avg_score / 2.0
            ax.text(
                x[i],
                mid_height,
                f"{avg_score:.3f}",
                ha="center",
                va="center",
                fontsize=6,
                color="black",
                clip_on=True
            )
            bottom += avg_score

        # Annotate total above each bar
        if bottom > 0:
            ax.text(
                x[i],
                bottom + (0.01 * bottom),
                f"sum={bottom:.3f}",
                ha="center",
                va="bottom",
                fontsize=8,
                rotation=90
            )

    # 6) Final formatting
    ax.set_xticks(x)
    ax.set_xticklabels(filtered_bio, rotation=90, fontsize=7)
    ax.set_ylabel("Average Score (stacked by disease)", fontsize=10)
    ax.set_xlabel("Biomarker", fontsize=10)
    ax.set_title(f"{category_name} Averages for Symptom: '{symptom}'", fontsize=12)

    # 7) Legend (colored patches for each disease)
    legend_handles = [
        Patch(facecolor=disease_to_color[d], edgecolor="black", label=d)
        for d in all_diseases
    ]
    ax.legend(
        handles=legend_handles,
        title="Disease",
        bbox_to_anchor=(1.02, 1),
        loc="upper left"
    )

    plt.tight_layout()
    plt.show()
