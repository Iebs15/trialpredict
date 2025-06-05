import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict
from cache_utiities import load_from_cache, save_to_cache, clean_old_caches

# ----------------------------------------------------------
# Configuration / Constants
# ----------------------------------------------------------


EXCEL_PATH = "processed_target_scores_colored_final_v5.xlsx"
GREEN_HEX = "C6EFCE"   # Inhibitor fill color
YELLOW_HEX = "FFEB9C"  # Promoter fill color

# ----------------------------------------------------------
# STEP 1: Load the entire workbook & DataFrame ONCE at startup
# ----------------------------------------------------------
print("⏳ Loading Excel into memory...")

# 1A. Load pandas DataFrame once
GLOBAL_DF = pd.read_excel(EXCEL_PATH)
GLOBAL_WB = load_workbook(EXCEL_PATH, data_only=True)

# 1B. Build a map: lowercase disease_name → list of DataFrame row indices
SYMP_ROW_MAP = defaultdict(list)
for idx, disease in enumerate(GLOBAL_DF["MappedDisease"]):
    if pd.notna(disease):
        SYMP_ROW_MAP[disease.strip().lower()].append(idx)

# 1C. Open the same workbook with openpyxl once
GLOBAL_WB = load_workbook(EXCEL_PATH, data_only=True)
GLOBAL_WS = GLOBAL_WB[GLOBAL_WB.sheetnames[0]]

# 1D. Build a mapping: column name → Excel column index
header_cells = GLOBAL_WS[1]
COL_NAME_TO_IDX = {cell.value: cell.col_idx for cell in header_cells if cell.value is not None}

# 1E. Pre‐compute a global fill‐color cache for every used cell
#     Keyed by (excel_row, excel_col) → 6‐digit RGB hex (or None)
print("⏳ Building GLOBAL_FILL_CACHE (cell‐color lookup)...")
GLOBAL_FILL_CACHE = {}
max_row = GLOBAL_WS.max_row
for row_idx in range(2, max_row + 1):  # skip header row 1
    for col_name, excel_col in COL_NAME_TO_IDX.items():
        cell = GLOBAL_WS.cell(row=row_idx, column=excel_col)
        if cell.fill and cell.fill.patternType == "solid" and cell.fill.start_color.rgb:
            # Take just the last 6 hex digits
            GLOBAL_FILL_CACHE[(row_idx, excel_col)] = cell.fill.start_color.rgb[-6:]
        else:
            GLOBAL_FILL_CACHE[(row_idx, excel_col)] = None

print("✅ Startup loading complete.\n")

# ----------------------------------------------------------
# STEP 2: process_disease_data now uses the preloaded data
# ----------------------------------------------------------
def process_disease_data(disease_name: str):
    """
    Returns a nested JSON‐like dict of biomarker→symptom→{percent_inhibitor, avg_inhibitor, ...}
    for all rows where MappedDisease == disease_name. Uses global caches to minimize latency.
    """
    # 2A) Purge old cache entries (> 2 days)
    clean_old_caches(days=2)

    # 2B) If we already computed this disease, return from cache
    cached_data = load_from_cache(disease_name)
    if cached_data is not None:
        print(f"🔄 Loaded data for '{disease_name}' from cache.")
        return cached_data

    # 2C) Find all DataFrame indices matching this disease (case‐insensitive)
    disease_key = disease_name.strip().lower()
    matching_indices = SYMP_ROW_MAP.get(disease_key, [])
    if not matching_indices:
        return {"error": f"No data found for disease: '{disease_name}'"}

    # 2D) Prepare list of biomarker columns (everything except id/symptom/disease)
    all_columns = list(GLOBAL_DF.columns)
    id_col = "drugId"
    symptom_col = "unique_symptom"
    mapped_col = "MappedDisease"
    biomarker_cols = [
        col for col in all_columns
        if col not in (id_col, symptom_col, mapped_col)
    ]

    # 2E) Build nested structure: biomarker → symptom → {'green': [vals], 'yellow': [vals]}
    data = defaultdict(lambda: defaultdict(lambda: {'green': [], 'yellow': []}))

    for df_idx in matching_indices:
        row = GLOBAL_DF.iloc[df_idx]
        symptom = row[symptom_col]
        excel_row = df_idx + 2  # DataFrame idx 0 maps to Excel row 2, etc.

        for biomarker in biomarker_cols:
            val = row[biomarker]
            if pd.isna(val):
                continue
            excel_col = COL_NAME_TO_IDX.get(biomarker)
            if not excel_col:
                continue
            color_hex = GLOBAL_FILL_CACHE.get((excel_row, excel_col))
            if color_hex == GREEN_HEX:
                data[biomarker][symptom]['green'].append(val)
            elif color_hex == YELLOW_HEX:
                data[biomarker][symptom]['yellow'].append(val)
            # else: either no fill or some other color → skip

    # 2F) Compute averages, percentages & build output JSON
    output = {}
    for biomarker, symptom_map in data.items():
        output[biomarker] = {}
        for symptom, vals in symptom_map.items():
            green_list  = vals['green']
            yellow_list = vals['yellow']

            avg_inh = sum(green_list)  / len(green_list)  if green_list  else 0.0
            avg_prm = sum(yellow_list) / len(yellow_list) if yellow_list else 0.0
            total   = avg_inh + avg_prm

            if total > 0:
                pct_inh = (avg_inh / total) * 100
                pct_prm = (avg_prm / total) * 100
            else:
                pct_inh = pct_prm = 0.0

            output[biomarker][symptom] = {
                "percent_inhibitor": round(pct_inh, 3),
                "avg_inhibitor":     round(avg_inh, 6),
                "avg_promoter":      round(avg_prm, 6),
                "percent_promoter":  round(pct_prm, 3),
                "total_avg":         round(total, 6),
            }

    # 2G) Cache the result and return
    save_to_cache(disease_name, output)
    return output
